#!/usr/bin/python
# -*- coding: UTF-8 -*-

import xlrd,xlwt
import openpyxl
from openpyxl.compat.numbers import NUMPY, PANDAS
from openpyxl.xml import DEFUSEDXML, LXML
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import openpyxl._constants as constants

def write_excel_xlrd():
    readbook = xlrd.open_workbook(r'test.xlsx')
    writebook = xlwt.Workbook()#打开一个excel
    sheet = writebook.add_sheet('test')#在打开的excel中添加一个sheet

    table = readbook.sheets()[0]#获取读入的文件的第一个sheet
    nrows = table.nrows#获取sheet的行数
    ncols = table.ncols

    for i in range(nrows):
        if i == 0:#我处理的数据第一行是属性名，所以去掉
            continue
        for j in range(ncols):
            ydh = table.cell(i,20).value
            if j == 20 and not ydh:
                sheet.write(i,j, 'aaa888')
            else:
                sheet.write(i,j,table.cell(i,j).value)

    writebook.save('bb.xls')#一定要记得保存

def write_excel_openpyxl():
    filename = r'test.xlsx'
    inwb = load_workbook(filename)#读文件
    outwb = Workbook()#打开一个将写的文件
    outws = outwb.create_sheet(title="cool")#在将写的文件创建sheet
    sheetnames = inwb.get_sheet_names()#获取读文件中所有的sheet，通过名字的方式
    ws = inwb.get_sheet_by_name(sheetnames[0])#获取第一个sheet内容
    rows = ws.max_row#获取读取的excel的文件的行数
    cols = ws.max_column
    for i in range(rows):
        if i<1:
            continue
        for j in range(cols):
            outws.cell(row=i+1, column=j+1).value = ws.cell(row=i+1, column=j+1).value
        ydh = ws.cell(row=i+1, column=21).value
        if not ydh:
            outws.cell(row = i+1,column=21).value = 'AAA123'
    outwb.save('aa.xlsx')


if __name__ == '__main__':
    #write_excel_openpyxl()
    write_excel_xlrd()

